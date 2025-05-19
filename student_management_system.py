import openpyxl
import random
import string

class StudentManagementSystem:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.workbook = self.load_workbook()
        self.sheet = self.workbook.active
        self.initialize_excel()
        self.students = self.load_students()

    # ... 其他方法 ...
    def initialize_excel(self):
        """
        初始化 Excel 文件，确保表头存在。

        如果表头不存在，直接在 Excel 表格的第一行写入表头。
        """
        header = ["学号", "姓名", "班级", "邮箱", "家庭住址", "科目1", "科目2", "科目3", "科目4", "科目5"]
        existing_header = self.sheet[1]
        if not existing_header or any(cell.value != header[i] for i, cell in enumerate(existing_header)):
            for i, value in enumerate(header, 1):
                self.sheet.cell(row=1, column=i, value=value)
            self.workbook.save(self.excel_file)


    def generate_random_info(self):
        student_id = ''.join(random.choices(string.digits, k=8))
        # 姓名生成
        surnames = ["吴", "王", "李", "张", "刘", "陈", "洪", "汪", "胡", "苏", "阮", "吕"]
        chinese_characters = "巧月巧云今夜巧非宁静无以致远超然离俗尘采菊东篱下低头思故乡岂同凡鸟群"
        given_name = random.choice(chinese_characters) + random.choice(chinese_characters)
        surname = random.choice(surnames)
        student_name = f"{surname}{given_name}".capitalize()

        student_class = random.choice(["1班", "2班", "3班"])
        student_email = f"{student_id}@qq.com"
        # 地址生成
        cities = ["合肥"]
        roads = ["安庆路", "黄山路", "潜山路", "长江路", "延乔路", "繁华大道", "望江路", "合作化南路", "金寨路",
                 "宁国路", "马鞍山路"]
        number = ''.join(random.choices(string.digits, k=2))
        student_address = f"{random.choice(cities)}{random.choice(roads)}{number} 号"
        subject_grades = [random.randint(60, 100) for _ in range(5)]
        student_info = {
            "学号": student_id,
            "姓名": student_name,
            "班级": student_class,
            "邮箱": student_email,
            "家庭住址": student_address,
            "科目1": subject_grades[0],
            "科目2": subject_grades[1],
            "科目3": subject_grades[2],
            "科目4": subject_grades[3],
            "科目5": subject_grades[4],
        }

        return student_info


    def load_students(self):
        students = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            student_info = {
                "学号": row[0],
                "姓名": row[1],
                "班级": row[2],
                "邮箱": row[3],
                "家庭住址": row[4],
                "科目1": row[5],
                "科目2": row[6],
                "科目3": row[7],
                "科目4": row[8],
                "科目5": row[9],
            }
            students.append(student_info)
        return students


    def load_workbook(self):
        try:
            return openpyxl.load_workbook(self.excel_file)
        except FileNotFoundError:
            return openpyxl.Workbook()


    def add_student(self, student_info):
        self.students.append(student_info)
        self.load_students()
        self.update_excel()

    def find_student(self, student_id):
        for student in self.students:
            if student["学号"] == student_id:
                return student
        return None

    def delete_student(self, student_id):
        for student in self.students:
            if student["学号"] == student_id:
                self.students.remove(student)
                self.update_excel()
                return True
        return False

    def update_student_info(self, student_id, new_info):
        for i, student in enumerate(self.students):
            if student["学号"] == student_id:
                self.students[i] = new_info
                self.update_excel()
                return True
        return False

    def update_excel(self):
        # 删除旧的学生信息行
        self.sheet.delete_rows(2, self.sheet.max_row)

        # 添加新的学生信息
        header = ["学号", "姓名", "班级", "邮箱", "家庭住址", "科目1", "科目2", "科目3", "科目4", "科目5"]
        # self.sheet.append(header)
        for student in self.students:
            row = [student[key] for key in header]
            self.sheet.append(row)

        # 保存更新后的 Excel 文件
        self.workbook.save(self.excel_file)




